"""World Bank Pink Sheet commodity price fetcher.

On-demand, session-cached commodity prices from World Bank Pink Sheet
(monthly Excel, free, no API key required).

Covers metals (copper, aluminum, zinc, nickel, lead) AND agricultural
commodities (corn, wheat, soybean). This is the RELIABLE replacement for:
- LME inventory (Cloudflare-blocked) → metal PRICES from WB
- USDA WASDE (PDF-only, unparseable) → grain PRICES from WB

Source: https://thedocs.worldbank.org (free, no key, monthly updates)

All data is CONTEXT only, not trading signals (Law 3 compliance).
Graceful degradation: return {"error": "source_unavailable"} on failure.
"""
from __future__ import annotations

import asyncio
import logging
import re
from datetime import datetime, timezone
from typing import Any

import httpx

from marketmind.integrity.input_guard import sanitize_for_llm_prompt

logger = logging.getLogger("marketmind.gateway.world_bank_fetcher")

# ---------------------------------------------------------------------------
# World Bank Pink Sheet constants
# ---------------------------------------------------------------------------
_WB_PINK_SHEET_URL = (
    "https://thedocs.worldbank.org/en/doc/"
    "5d903e848db1d1b83e0ec8f744e55570-0350012021/related/"
    "CMO-Historical-Data-Monthly.xlsx"
)

# Column name patterns for metals and grains in the Pink Sheet Excel
_WB_METAL_COLUMNS = {
    "copper": ["Copper", "copper", "COPPER"],
    "aluminum": ["Aluminum", "aluminium", "ALUMINUM", "Aluminium"],
    "zinc": ["Zinc", "zinc", "ZINC"],
    "nickel": ["Nickel", "nickel", "NICKEL"],
    "lead": ["Lead", "lead", "LEAD"],
}
_WB_GRAIN_COLUMNS = {
    "corn": ["Maize", "Corn", "corn", "MAIZE", "CORN"],
    "wheat": ["Wheat", "wheat", "WHEAT"],
    "soybean": ["Soybean", "Soybeans", "soybean", "SOYBEAN"],
}

# ---------------------------------------------------------------------------
# Session-level cache
# ---------------------------------------------------------------------------
_cache: dict[str, dict] = {}
_cache_locks: dict[str, asyncio.Lock] = {}


def _clear_cache() -> None:
    """Clear the module-level cache (used between tests)."""
    _cache.clear()
    _cache_locks.clear()


def _sanitize(data: dict, source: str = "commodity_data") -> dict:
    """Sanitize all string fields in a data dict before LLM consumption."""
    for key, val in list(data.items()):
        if isinstance(val, str):
            result = sanitize_for_llm_prompt(val, source=source)
            data[key] = result.sanitized
            for warning in result.warnings:
                logger.warning("input_guard [%s] field=%s: %s", source, key, warning)
    return data


def _parse_float(val: Any) -> float:
    """Parse a value to float, returning 0.0 on failure."""
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def _redact(text: str) -> str:
    """Replace api_key=XXXX with api_key=*** to prevent key leakage in logs."""
    return re.sub(r'api_key=[^&\s\'\"<>]+', 'api_key=***', text)


# ===================================================================
# Public API
# ===================================================================


async def get_world_bank_commodities() -> dict:
    """Fetch commodity prices from World Bank Pink Sheet (monthly Excel, free).

    Covers metals (copper, aluminum, zinc, nickel, lead) AND agricultural
    commodities (corn, wheat, soybean). This is the RELIABLE replacement for:
    - LME inventory (Cloudflare-blocked) → metal PRICES from WB
    - USDA WASDE (PDF-only, unparseable) → grain PRICES from WB

    Source: https://thedocs.worldbank.org (free, no key, monthly updates)
    """
    key = "wb_commodities"
    if key in _cache:
        return _cache[key]
    if key not in _cache_locks:
        _cache_locks[key] = asyncio.Lock()
    async with _cache_locks[key]:
        if key in _cache:
            return _cache[key]
        result = await _fetch_world_bank()
        _cache[key] = result
        return result


# ===================================================================
# World Bank Pink Sheet implementation
# ===================================================================


async def _fetch_world_bank() -> dict:
    """Fetch latest commodity prices from World Bank Pink Sheet monthly Excel."""
    client = httpx.AsyncClient(timeout=httpx.Timeout(50.0))
    try:
        resp = await client.get(_WB_PINK_SHEET_URL)
        resp.raise_for_status()

        import io
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)
        except ImportError:
            logger.warning("openpyxl not available — cannot parse World Bank Pink Sheet")
            return _sanitize({
                "error": "source_unavailable",
                "detail": "openpyxl not installed. Run: pip install openpyxl",
            }, "wb_data")

        sheet = wb["Monthly Prices"] if "Monthly Prices" in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))

        if len(rows) < 3:
            return _sanitize({"error": "source_unavailable", "detail": "Pink Sheet too short"}, "wb_data")

        # Header row (usually row 3-5)
        headers = [str(h).strip() if h else "" for h in (rows[3] or [])]

        # Find metal and grain columns
        def find_column(patterns):
            for i, h in enumerate(headers):
                for p in patterns:
                    if p.lower() in h.lower():
                        return i
            return None

        metals = {}
        for name, patterns in _WB_METAL_COLUMNS.items():
            col = find_column(patterns)
            if col is not None:
                # Get last row's value
                for row in reversed(rows[4:]):
                    if row and col < len(row) and row[col] is not None:
                        try:
                            metals[name] = float(row[col])
                        except (ValueError, TypeError):
                            pass
                        break

        grains = {}
        for name, patterns in _WB_GRAIN_COLUMNS.items():
            col = find_column(patterns)
            if col is not None:
                for row in reversed(rows[4:]):
                    if row and col < len(row) and row[col] is not None:
                        try:
                            grains[name] = float(row[col])
                        except (ValueError, TypeError):
                            pass
                        break

        wb.close()

        if not metals and not grains:
            return _sanitize({
                "error": "source_unavailable",
                "detail": "Could not locate metal or grain columns in Pink Sheet",
            }, "wb_data")

        return _sanitize({
            "indicator": "world_bank_commodities",
            "metals": metals,
            "grains": grains,
            "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "source": "world_bank",
            "cadence": "monthly",
        }, "wb_data")

    except httpx.HTTPStatusError as e:
        logger.warning("World Bank Pink Sheet HTTP error: %s", e)
        return _sanitize({
            "error": "source_unavailable",
            "detail": f"World Bank returned HTTP {e.response.status_code}",
        }, "wb_data")
    except Exception as e:
        logger.warning("World Bank fetch failed: %s", e)
        return _sanitize({
            "error": "source_unavailable",
            "detail": _redact(str(e)),
        }, "wb_data")
    finally:
        await client.aclose()
