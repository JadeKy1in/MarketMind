"""On-demand, session-cached commodity supply/demand data fetchers.

Three public async functions, all using direct httpx (no new dependencies):

- get_lme_metal_inventory(metal) — LME warehouse stocks (public JSON API, no key)
- get_usda_wasde() — USDA WASDE monthly supply/demand (graceful degradation)
- get_eia_extended(product) — EIA extended data (API key: EIA_KEY)

All data is CONTEXT only, not trading signals (Law 3 compliance).
Session-level in-memory cache (same pattern as macro_data.py).
Graceful degradation: return {"error": "source_unavailable"} on failure.
"""
from __future__ import annotations

import asyncio
import logging
import os
import re
from datetime import datetime, timezone
from typing import Any

import httpx

from marketmind.integrity.input_guard import sanitize_for_llm_prompt

logger = logging.getLogger("marketmind.gateway.commodity_fetcher")

# ---------------------------------------------------------------------------
# LME API constants
# ---------------------------------------------------------------------------
_LME_BASE = "https://www.lme.com/api/price-month"

_LME_METALS: dict[str, dict[str, str]] = {
    "copper":    {"metal_key": "copper",    "label": "LME Copper Grade A",  "cadence": "daily"},
    "aluminum":  {"metal_key": "aluminium", "label": "LME Aluminium",       "cadence": "daily"},
    "zinc":      {"metal_key": "zinc",      "label": "LME Zinc",            "cadence": "daily"},
    "nickel":    {"metal_key": "nickel",    "label": "LME Nickel",          "cadence": "daily"},
    "lead":      {"metal_key": "lead",      "label": "LME Lead",            "cadence": "daily"},
}

# ---------------------------------------------------------------------------
# USDA WASDE constants
# ---------------------------------------------------------------------------
_USDA_BASE = "https://www.usda.gov/oce/commodity/wasde"

# ---------------------------------------------------------------------------
# EIA extended constants
# ---------------------------------------------------------------------------
_EIA_BASE = "https://api.eia.gov/v2"

_EIA_EXTENDED: dict[str, dict[str, str]] = {
    "natgas_storage": {
        "route": "/natural-gas/stor/wkly/data/",
        "label": "Natural Gas Weekly Storage",
        "cadence": "weekly",
    },
    "crude_production": {
        "route": "/petroleum/crd/crdp/data/",
        "label": "Crude Oil Production",
        "cadence": "monthly",
    },
}

# ---------------------------------------------------------------------------
# Session-level cache
# ---------------------------------------------------------------------------
_cache: dict[str, dict] = {}
_cache_locks: dict[str, asyncio.Lock] = {}


def _clear_cache() -> None:
    """Clear the module-level cache (used between tests)."""
    _cache.clear()
    _cache_locks.clear()


def _sanitize(data: dict, source: str = "commodity_data") -> dict:
    """Sanitize all string fields in a data dict before LLM consumption."""
    for key, val in list(data.items()):
        if isinstance(val, str):
            result = sanitize_for_llm_prompt(val, source=source)
            data[key] = result.sanitized
            for warning in result.warnings:
                logger.warning("input_guard [%s] field=%s: %s", source, key, warning)
    return data


# ===================================================================
# Public API
# ===================================================================


async def get_lme_metal_inventory(metal: str) -> dict:
    """Fetch metal price — tries LME JSON first, falls back to World Bank.

    LME is Cloudflare-protected (often 403). World Bank Pink Sheet provides
    monthly metal prices (free, no key) as reliable fallback.
    """
    metal = metal.strip().lower()
    key = f"lme:{metal}"
    if key in _cache:
        return _cache[key]
    if key not in _cache_locks:
        _cache_locks[key] = asyncio.Lock()
    async with _cache_locks[key]:
        if key in _cache:
            return _cache[key]

        # Try LME JSON first
        result = await _fetch_lme(metal)
        if "error" not in result:
            _cache[key] = result
            return result

        # Fallback: World Bank Pink Sheet
        wb = await get_world_bank_commodities()
        if "error" not in wb and metal in wb.get("metals", {}):
            result = {
                "metal": metal,
                "label": _LME_METALS.get(metal, {}).get("label", metal.title()),
                "price_usd": wb["metals"][metal],
                "date": wb["date"],
                "source": "world_bank",
                "cadence": "monthly",
            }
        _cache[key] = result
        return result


async def get_world_bank_commodities() -> dict:
    """Fetch commodity prices from World Bank Pink Sheet (monthly Excel, free).

    Covers metals (copper, aluminum, zinc, nickel, lead) AND agricultural
    commodities (corn, wheat, soybean). This is the RELIABLE replacement for:
    - LME inventory (Cloudflare-blocked) → metal PRICES from WB
    - USDA WASDE (PDF-only, unparseable) → grain PRICES from WB

    Source: https://thedocs.worldbank.org (free, no key, monthly updates)
    """
    key = "wb_commodities"
    if key in _cache:
        return _cache[key]
    if key not in _cache_locks:
        _cache_locks[key] = asyncio.Lock()
    async with _cache_locks[key]:
        if key in _cache:
            return _cache[key]
        result = await _fetch_world_bank()
        _cache[key] = result
        return result


async def get_usda_wasde() -> dict:
    """Legacy stub — WASDE is PDF-only. Use get_world_bank_commodities() instead."""
    return await get_world_bank_commodities()


async def get_eia_extended(product: str) -> dict:
    """Fetch extended EIA data beyond the basic petroleum inventory.

    Args:
        product: "natgas_storage" or "crude_production".
    """
    product = product.strip().lower()
    key = f"eia_ext:{product}"
    if key in _cache:
        return _cache[key]
    if key not in _cache_locks:
        _cache_locks[key] = asyncio.Lock()
    async with _cache_locks[key]:
        if key in _cache:
            return _cache[key]
        result = await _fetch_eia_extended(product)
        _cache[key] = result
        return result


# ===================================================================
# LME implementation
# ===================================================================


async def _fetch_lme(metal: str) -> dict:
    """Fetch LME warehouse stocks for a metal."""
    metal_info = _LME_METALS.get(metal)
    if metal_info is None:
        return _sanitize({
            "error": "source_unavailable",
            "detail": f"Unknown metal: '{metal}'. Supported: copper, aluminum, zinc, nickel, lead",
        }, "lme_data")

    now_utc = datetime.now(timezone.utc)
    month_str = now_utc.strftime("%Y-%m")
    metal_key = metal_info["metal_key"]
    url = f"{_LME_BASE}?metal={metal_key}&month={month_str}"

    client = httpx.AsyncClient(timeout=httpx.Timeout(10.0))
    try:
        resp = await client.get(url)
        resp.raise_for_status()
        data = resp.json()
        inventory = _extract_lme_inventory(data)

        if inventory is None:
            return _sanitize({
                "error": "source_unavailable",
                "detail": f"LME response did not contain inventory data for {metal}",
            }, "lme_data")

        return _sanitize({
            "metal": metal,
            "label": metal_info["label"],
            "inventory_tonnes": inventory,
            "date": now_utc.strftime("%Y-%m-%d"),
            "source": "lme",
            "cadence": metal_info["cadence"],
        }, "lme_data")

    except httpx.HTTPStatusError as e:
        logger.warning("LME API error for %s: %s", metal, _redact(str(e)))
        return _sanitize({
            "error": "source_unavailable",
            "detail": f"LME API returned HTTP {e.response.status_code}",
        }, "lme_data")
    except Exception as e:
        logger.warning("LME fetch failed for %s: %s", metal, _redact(str(e)))
        return _sanitize({
            "error": "source_unavailable",
            "detail": _redact(str(e)),
        }, "lme_data")
    finally:
        await client.aclose()


def _extract_lme_inventory(data: dict | list) -> float | None:
    """Extract warehouse inventory tonnage from LME JSON response.

    Handles 4 JSON shapes: direct keys, nested dicts, list-wrapped, first-element.
    """
    if isinstance(data, list):
        data = data[0] if data else {}
    if not isinstance(data, dict):
        return None

    # Direct inventory fields
    for field in ("warehouse_stocks", "inventory", "stocks", "total_stocks",
                  "on_warrant", "live_warrants"):
        if field in data:
            val = _parse_float(data[field])
            if val > 0:
                return val

    # Nested under data/summary/warehouse
    for sub_key in ("data", "summary", "warehouse"):
        sub = data.get(sub_key)
        if isinstance(sub, dict):
            for inv_key in ("inventory", "stocks", "warehouse_stocks", "total", "tonnes"):
                if inv_key in sub:
                    val = _parse_float(sub[inv_key])
                    if val > 0:
                        return val

    # Nested list: check first element
    for list_key in ("data", "items", "records", "warehouses"):
        items = data.get(list_key)
        if isinstance(items, list) and items:
            first = items[0]
            if isinstance(first, dict):
                for inv_key in ("inventory", "stocks", "tonnes", "warehouse_stocks"):
                    if inv_key in first:
                        val = _parse_float(first[inv_key])
                        if val > 0:
                            return val

    return None


# ===================================================================
# USDA WASDE implementation
# ===================================================================


async def _fetch_wasde() -> dict:
    """Attempt to fetch USDA WASDE data; degrade gracefully on failure."""
    client = httpx.AsyncClient(timeout=httpx.Timeout(10.0))
    try:
        resp = await client.get(_USDA_BASE)
        resp.raise_for_status()
        html = resp.text

        corn = _extract_wasde_value(html, "corn")
        soy = _extract_wasde_value(html, "soybean")
        wheat = _extract_wasde_value(html, "wheat")

        if corn is None and soy is None and wheat is None:
            return _sanitize({
                "error": "source_unavailable",
                "detail": "USDA WASDE data not available in machine-readable format",
            }, "usda_data")

        date_val = _extract_wasde_date(html)
        return _sanitize({
            "report": "wasde",
            "corn_ending_stocks": corn or 0,
            "soybean_ending_stocks": soy or 0,
            "wheat_ending_stocks": wheat or 0,
            "date": date_val,
            "source": "usda",
            "cadence": "monthly",
        }, "usda_data")

    except httpx.HTTPStatusError as e:
        logger.warning("USDA WASDE HTTP error: %s", e)
        return _sanitize({
            "error": "source_unavailable",
            "detail": f"USDA WASDE page returned HTTP {e.response.status_code}",
        }, "usda_data")
    except Exception as e:
        logger.warning("USDA WASDE fetch failed: %s", e)
        return _sanitize({
            "error": "source_unavailable",
            "detail": str(e),
        }, "usda_data")
    finally:
        await client.aclose()


def _extract_wasde_value(html: str, commodity: str) -> float | None:
    """Search HTML for a WASDE ending stocks figure (million bushels)."""
    patterns = [
        re.escape(commodity) + r'[^>]{0,300}ending\s+stocks[^>]{0,100}?(\d{1,3}(?:,\d{3})*(?:\.\d+)?)',
        r'ending\s+stocks[^>]{0,300}' + re.escape(commodity) + r'[^>]{0,100}?(\d{1,3}(?:,\d{3})*(?:\.\d+)?)',
    ]
    for pat in patterns:
        m = re.search(pat, html, re.IGNORECASE)
        if m:
            try:
                return float(m.group(1).replace(",", ""))
            except ValueError:
                continue
    return None


def _extract_wasde_date(html: str) -> str:
    """Extract WASDE report date from the web page."""
    date_patterns = [
        r'WASDE\s+(?:report\s+)?[–-]\s*(\w+\s+\d{1,2},\s+\d{4})',
        r'(\w+\s+\d{1,2},\s+\d{4})\s+WASDE',
        r'Released\s+(\w+\s+\d{1,2},\s+\d{4})',
        r'(\d{4}-\d{2}-\d{2})',
    ]
    for pat in date_patterns:
        m = re.search(pat, html, re.IGNORECASE)
        if m:
            return m.group(1)
    return datetime.now(timezone.utc).strftime("%Y-%m-%d")


# ===================================================================
# World Bank Pink Sheet implementation (replaces LME + USDA)
# ===================================================================

_WB_PINK_SHEET_URL = (
    "https://thedocs.worldbank.org/en/doc/"
    "5d903e848db1d1b83e0ec8f744e55570-0350012021/related/"
    "CMO-Historical-Data-Monthly.xlsx"
)

# Column name patterns for metals and grains in the Pink Sheet Excel
_WB_METAL_COLUMNS = {
    "copper": ["Copper", "copper", "COPPER"],
    "aluminum": ["Aluminum", "aluminium", "ALUMINUM", "Aluminium"],
    "zinc": ["Zinc", "zinc", "ZINC"],
    "nickel": ["Nickel", "nickel", "NICKEL"],
    "lead": ["Lead", "lead", "LEAD"],
}
_WB_GRAIN_COLUMNS = {
    "corn": ["Maize", "Corn", "corn", "MAIZE", "CORN"],
    "wheat": ["Wheat", "wheat", "WHEAT"],
    "soybean": ["Soybean", "Soybeans", "soybean", "SOYBEAN"],
}


async def _fetch_world_bank() -> dict:
    """Fetch latest commodity prices from World Bank Pink Sheet monthly Excel."""
    client = httpx.AsyncClient(timeout=httpx.Timeout(50.0))
    try:
        resp = await client.get(_WB_PINK_SHEET_URL)
        resp.raise_for_status()

        import io
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)
        except ImportError:
            logger.warning("openpyxl not available — cannot parse World Bank Pink Sheet")
            return _sanitize({
                "error": "source_unavailable",
                "detail": "openpyxl not installed. Run: pip install openpyxl",
            }, "wb_data")

        sheet = wb["Monthly Prices"] if "Monthly Prices" in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))

        if len(rows) < 3:
            return _sanitize({"error": "source_unavailable", "detail": "Pink Sheet too short"}, "wb_data")

        # Header row (usually row 3-5)
        headers = [str(h).strip() if h else "" for h in (rows[3] or [])]

        # Find metal and grain columns
        def find_column(patterns):
            for i, h in enumerate(headers):
                for p in patterns:
                    if p.lower() in h.lower():
                        return i
            return None

        metals = {}
        for name, patterns in _WB_METAL_COLUMNS.items():
            col = find_column(patterns)
            if col is not None:
                # Get last row's value
                for row in reversed(rows[4:]):
                    if row and col < len(row) and row[col] is not None:
                        try:
                            metals[name] = float(row[col])
                        except (ValueError, TypeError):
                            pass
                        break

        grains = {}
        for name, patterns in _WB_GRAIN_COLUMNS.items():
            col = find_column(patterns)
            if col is not None:
                for row in reversed(rows[4:]):
                    if row and col < len(row) and row[col] is not None:
                        try:
                            grains[name] = float(row[col])
                        except (ValueError, TypeError):
                            pass
                        break

        wb.close()

        if not metals and not grains:
            return _sanitize({
                "error": "source_unavailable",
                "detail": "Could not locate metal or grain columns in Pink Sheet",
            }, "wb_data")

        return _sanitize({
            "indicator": "world_bank_commodities",
            "metals": metals,
            "grains": grains,
            "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "source": "world_bank",
            "cadence": "monthly",
        }, "wb_data")

    except httpx.HTTPStatusError as e:
        logger.warning("World Bank Pink Sheet HTTP error: %s", e)
        return _sanitize({
            "error": "source_unavailable",
            "detail": f"World Bank returned HTTP {e.response.status_code}",
        }, "wb_data")
    except Exception as e:
        logger.warning("World Bank fetch failed: %s", e)
        return _sanitize({
            "error": "source_unavailable",
            "detail": _redact(str(e)),
        }, "wb_data")
    finally:
        await client.aclose()


# EIA extended implementation
# ===================================================================


async def _fetch_eia_extended(product: str) -> dict:
    """Fetch extended EIA data (natgas storage, crude production)."""
    product_info = _EIA_EXTENDED.get(product)
    if product_info is None:
        return _sanitize({
            "error": "source_unavailable",
            "detail": f"Unknown product: '{product}'. Supported: natgas_storage, crude_production",
        }, "eia_data")

    eia_key = _get_eia_key()
    if not eia_key:
        return _sanitize({
            "error": "source_unavailable",
            "detail": "EIA_KEY not configured",
        }, "eia_data")

    route = product_info["route"]
    url = (
        f"{_EIA_BASE}{route}"
        f"?api_key={eia_key}"
        f"&frequency={product_info['cadence']}"
        f"&data[0]=value"
        f"&sort[0][column]=period"
        f"&sort[0][direction]=desc"
        f"&length=1"
    )

    client = httpx.AsyncClient(timeout=httpx.Timeout(10.0))
    try:
        resp = await client.get(url)
        resp.raise_for_status()
        data = resp.json()
        response_data = data.get("response", {})
        records = response_data.get("data", [])

        if not records:
            return _sanitize({
                "error": "source_unavailable",
                "detail": f"No EIA data for {product}",
            }, "eia_data")

        record = records[0]
        value = _parse_float(record.get("value", 0))
        unit = record.get("units", "")

        return _sanitize({
            "product": product,
            "label": product_info["label"],
            "value": value,
            "unit": unit,
            "date": record.get("period", ""),
            "source": "eia",
            "cadence": product_info["cadence"],
        }, "eia_data")

    except httpx.HTTPStatusError as e:
        logger.warning("EIA extended API error for %s: %s", product, _redact(str(e)))
        return _sanitize({
            "error": "source_unavailable",
            "detail": f"EIA API returned HTTP {e.response.status_code}",
        }, "eia_data")
    except Exception as e:
        logger.warning("EIA extended fetch failed for %s: %s", product, _redact(str(e)))
        return _sanitize({
            "error": "source_unavailable",
            "detail": _redact(str(e)),
        }, "eia_data")
    finally:
        await client.aclose()


# ===================================================================
# Helpers
# ===================================================================


def _get_eia_key() -> str:
    """Get EIA API key from config or environment."""
    try:
        from marketmind.config.settings import MarketMindConfig
        cfg = MarketMindConfig()
        if hasattr(cfg, "eia_key") and cfg.eia_key:
            return cfg.eia_key
    except (ImportError, AttributeError) as e:
        logger.warning("EIA API key not available from MarketMindConfig: %s", e)
    return os.environ.get("EIA_KEY", "").strip()


def _parse_float(val: Any) -> float:
    """Parse a value to float, returning 0.0 on failure."""
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def _redact(text: str) -> str:
    """Replace api_key=XXXX with api_key=*** to prevent key leakage in logs."""
    return re.sub(r'api_key=[^&\s\'\"<>]+', 'api_key=***', text)
